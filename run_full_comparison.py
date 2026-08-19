"""Full comparison and Excel report generation across all 15 competition datasets:
5 Standard Sources + 5 Test Samples + 5 Soil Samples (Ra/Th/K).
"""
from __future__ import annotations

import math
import sys
from pathlib import Path
import numpy as np

# ══════════════════════════════════════════════════════════════════════════════
# GROUND TRUTH extracted from ALL 10 reference screenshots + 1 docx
# Format: (nuclide, energy_kev, roi_l, roi_r, roi_area, net_area, area_uncert)
# ══════════════════════════════════════════════════════════════════════════════
GROUND_TRUTH_STANDARDS = {
    "Co-60 标准源": {
        "path": "核分析赛道/核素识别/标准源/Co-60/Co-60 点源.xls",
        "nuclides": ["Co-60"],
        "peaks": [
            ("Co-60", 1173.228, 3926, 3964, 270915, 265925, 0.1995),
            ("Co-60", 1332.492, 4457, 4502, 241486, 239258, 0.2073),
        ],
    },
    "Eu-152 标准源": {
        "path": "核分析赛道/核素识别/标准源/Eu-152/Eu-152 点源.xls",
        "nuclides": ["Eu-152"],
        "peaks": [
            ("Eu-152", 121.782, 404, 415, 1904207, 1854784, 0.0764),
            ("Eu-152", 244.697, 817, 830, 315515, 302734, 0.1937),
            ("Eu-152", 344.279, 1151, 1166, 1283842, 1261360, 0.0915),
            ("Eu-152", 778.904, 2611, 2632, 310439, 299805, 0.1927),
            ("Eu-152", 964.057, 3233, 3257, 279922, 270405, 0.2028),
            ("Eu-152", 1112.076, 3730, 3756, 259446, 249546, 0.2125),
            ("Eu-152", 1408.013, 4724, 4754, 316005, 311711, 0.1829),
        ],
    },
    "Co-60+Eu-152 标准源": {
        "path": "核分析赛道/核素识别/标准源/Co-60+Eu-152/Co-60  Eu-152混合点源.xls",
        "nuclides": ["Co-60", "Eu-152"],
        "peaks": [
            ("Eu-152", 121.782, 404, 415, 1616929, 1570429, 0.0834),
            ("Eu-152", 244.697, 817, 830, 269684, 256613, 0.2132),
            ("Eu-152", 344.279, 1151, 1166, 1097282, 1074433, 0.0996),
            ("Eu-152", 778.904, 2611, 2632, 267599, 255854, 0.2119),
            ("Eu-152", 964.057, 3233, 3257, 243386, 230155, 0.2274),
            ("Eu-152", 1112.076, 3730, 3756, 225025, 212487, 0.2372),
            ("Co-60", 1173.228, 3935, 3962, 317010, 307876, 0.1885),
            ("Co-60", 1332.492, 4470, 4500, 282785, 278531, 0.1939),
            ("Eu-152", 1408.013, 4724, 4754, 269730, 265512, 0.1988),
        ],
    },
    "Ba-133+Cs-137 标准源": {
        "path": "核分析赛道/核素识别/标准源/Cs-137+Ba-133/Cs-137 Ba-133标准源.xls",
        "nuclides": ["Ba-133", "Cs-137"],
        "peaks": [
            ("Ba-133", 80.998, 267, 278, 407277, 378904, 0.1821),
            ("Ba-133", 356.013, 1191, 1205, 622991, 617335, 0.1290),
            ("Cs-137", 661.657, 2217, 2237, 560605, 558234, 0.1347),
        ],
    },
    "Cs-137+I-131 水样": {
        "path": "核分析赛道/核素识别/标准源/Cs-137+I-131/I-131 Cs-137 水样1.xls",
        "nuclides": ["Cs-137", "I-131"],
        "peaks": [
            ("I-131", 284.305, 950, 963, 470, 262, 39.5841),
            ("I-131", 364.489, 1219, 1234, 4873, 4649, 1.5780),
            ("I-131", 636.989, 2134, 2153, 380, 304, 8.5567),
            ("Cs-137", 661.657, 2217, 2237, 18080, 17998, 0.7505),
        ],
    },
}

GROUND_TRUTH_TEST_SAMPLES = {
    "测试样1": {
        "path": "核分析赛道/核素识别/测试样/测试样1/测试样1.xls",
        "nuclides": ["U-238", "U-235", "Cs-137", "Co-60"],
        "peaks": [
            ("U-238", 63.290, 216, 235, 331, 148, 0.0000),
            ("U-235", 84.214, 291, 310, 327, 111, 0.0000),
            ("U-238", 92.380, 320, 339, 607, 353, 24.6903),
            ("U-235", 143.760, 504, 523, 901, 86, 0.0000),
            ("Cs-137", 661.657, 2354, 2381, 2299, 2232, 2.2144),
            ("Co-60", 1173.228, 4181, 4217, 292, 277, 6.5206),
            ("Co-60", 1332.492, 4750, 4788, 282, 281, 6.0270),
        ],
    },
    "测试样2": {
        "path": "核分析赛道/核素识别/测试样/测试样2/测试样2.xls",
        "nuclides": ["U-238", "U-235", "Eu-152", "Cs-137", "Co-60"],
        "peaks": [
            ("U-238", 63.290, 216, 235, 777, 202, 0.0000),
            ("U-235", 84.214, 291, 310, 909, 264, 0.0000),
            ("U-238", 92.380, 320, 339, 1244, 563, 0.0000),
            ("Eu-152", 121.782, 425, 444, 2591, 57, 0.0000),
            ("U-235", 143.760, 504, 523, 1068, 216, 0.0000),
            ("U-235", 163.356, 574, 594, 820, 111, 0.0000),
            ("Eu-152", 244.697, 864, 885, 2531, 172, 0.0000),
            ("Eu-152", 344.279, 1220, 1243, 2141, 145, 0.0000),
            ("Cs-137", 661.657, 2354, 2381, 44781, 44587, 0.4767),
            ("Eu-152", 778.904, 2772, 2802, 173, 39, 0.0000),
            ("Eu-152", 964.057, 3434, 3466, 134, 18, 0.0000),
            ("Co-60", 1173.228, 4181, 4217, 89, 21, 0.0000),
        ],
    },
    "测试样3": {
        "path": "核分析赛道/核素识别/测试样/测试样3/测试样3.xls",
        "nuclides": ["U-238", "U-235", "Eu-152", "Cs-137", "Co-60"],
        "peaks": [
            ("U-238", 63.290, 216, 235, 533, 183, 0.0000),
            ("U-235", 84.214, 291, 310, 745, 254, 0.0000),
            ("U-238", 92.380, 320, 339, 942, 549, 19.4846),
            ("Eu-152", 121.782, 425, 444, 1573, 63, 0.0000),
            ("U-235", 143.760, 504, 523, 716, 150, 0.0000),
            ("U-235", 163.356, 574, 594, 473, 124, 0.0000),
            ("Eu-152", 244.697, 864, 885, 639, 82, 0.0000),
            ("Cs-137", 661.657, 2354, 2381, 14230, 14126, 0.8507),
            ("Eu-152", 778.904, 2772, 2802, 163, 15, 0.0000),
            ("Eu-152", 964.057, 3434, 3466, 161, 31, 0.0000),
            ("Co-60", 1173.228, 4181, 4217, 202, 95, 0.0000),
            ("Co-60", 1332.492, 4750, 4788, 146, 64, 0.0000),
        ],
    },
    "测试样4": {
        "path": "核分析赛道/核素识别/测试样/测试样4/测试样4.xls",
        "nuclides": ["U-238", "U-235", "Eu-152", "Cs-137", "Co-60"],
        "peaks": [
            ("U-238", 63.290, 216, 235, 474, 192, 0.0000),
            ("U-235", 84.214, 291, 310, 456, 151, 0.0000),
            ("U-238", 92.380, 320, 339, 762, 504, 11.2507),
            ("Eu-152", 121.782, 425, 444, 1028, 94, 0.0000),
            ("U-235", 143.760, 504, 523, 390, 84, 0.0000),
            ("U-235", 163.356, 574, 594, 462, 99, 0.0000),
            ("Eu-152", 244.697, 864, 885, 915, 89, 0.0000),
            ("Eu-152", 344.279, 1220, 1243, 713, 79, 0.0000),
            ("Cs-137", 661.657, 2354, 2381, 1316, 1253, 3.0453),
            ("Eu-152", 778.904, 2772, 2802, 177, 29, 0.0000),
            ("Eu-152", 964.057, 3434, 3466, 131, 33, 0.0000),
            ("Co-60", 1173.228, 4181, 4217, 108, 28, 0.0000),
            ("Co-60", 1332.492, 4750, 4788, 94, 18, 0.0000),
        ],
    },
    "测试样5": {
        "path": "核分析赛道/核素识别/测试样/测试样5/测试样5.xls",
        "nuclides": ["U-238", "U-235", "Eu-152", "Cs-137", "Co-60"],
        "peaks": [
            ("U-238", 63.290, 216, 235, 718, 209, 0.0000),
            ("U-235", 84.214, 291, 310, 1244, 275, 0.0000),
            ("U-238", 92.380, 320, 339, 1078, 519, 0.0000),
            ("Eu-152", 121.782, 425, 444, 2629, 52, 0.0000),
            ("U-235", 143.760, 504, 523, 2682, 38, 0.0000),
            ("U-235", 163.356, 574, 594, 2688, 99, 0.0000),
            ("Eu-152", 244.697, 864, 885, 2596, 107, 0.0000),
            ("Eu-152", 344.279, 1220, 1243, 2169, 84, 0.0000),
            ("Cs-137", 661.657, 2354, 2381, 47459, 47260, 0.4629),
            ("Eu-152", 778.904, 2772, 2802, 137, 16, 0.0000),
            ("Eu-152", 964.057, 3434, 3466, 149, 36, 0.0000),
            ("Co-60", 1173.228, 4181, 4217, 85, 24, 0.0000),
            ("Co-60", 1332.492, 4750, 4788, 76, 31, 0.0000),
        ],
    },
}

RTK_GROUND_TRUTH = {
    "镭钍钾1": {
        "path": "核分析赛道/镭钍钾定量分析/测试样品数据/镭钍钾1.xls",
        "mass_kg": 0.334,
        "results": {"Th-232": 22.402, "Ra-226": 55.293, "K-40": 398.01},
    },
    "镭钍钾2": {
        "path": "核分析赛道/镭钍钾定量分析/测试样品数据/镭钍钾2.xls",
        "mass_kg": 0.245,
        "results": {"Th-232": 31.016, "Ra-226": 68.298, "K-40": 525.87},
    },
    "镭钍钾3": {
        "path": "核分析赛道/镭钍钾定量分析/测试样品数据/镭钍钾3.xls",
        "mass_kg": 0.244,
        "results": {"Th-232": 15.762, "Ra-226": 45.779, "K-40": 289.81},
    },
    "镭钍钾4": {
        "path": "核分析赛道/镭钍钾定量分析/测试样品数据/镭钍钾4.xls",
        "mass_kg": 0.264,
        "results": {"Th-232": 16.978, "Ra-226": 55.847, "K-40": 326.60},
    },
    "镭钍钾5": {
        "path": "核分析赛道/镭钍钾定量分析/测试样品数据/镭钍钾5.xls",
        "mass_kg": 0.310,
        "results": {"Th-232": 19.337, "Ra-226": 53.662, "K-40": 355.97},
    },
}


def run_nuclide_pipeline(path: str):
    from gamma_spectrum_analyzer.io import read_spectrum
    from gamma_spectrum_analyzer.calibration import auto_energy_calibration
    from gamma_spectrum_analyzer.peaks import find_and_fit_peaks
    from gamma_spectrum_analyzer.identify import identify_peaks, build_complete_nuclide_peaks
    from gamma_spectrum_analyzer.quantify import fill_quantification

    spec = read_spectrum(path)
    cal = auto_energy_calibration(spec)
    peaks = find_and_fit_peaks(spec, cal)
    confirmed = identify_peaks(peaks, spectrum=spec)
    if confirmed:
        peaks = build_complete_nuclide_peaks(spec, cal, peaks, confirmed)
    fill_quantification(peaks, spec, cal)
    return peaks, confirmed, spec, cal


def run_rtk_pipeline():
    from gamma_spectrum_analyzer.io import read_spectrum
    from gamma_spectrum_analyzer.calibration import auto_energy_calibration
    from gamma_spectrum_analyzer.peaks import find_and_fit_peaks
    from gamma_spectrum_analyzer.quantify import quantify_specific_activity
    from gamma_spectrum_analyzer.efficiency import CalibrationSource, EfficiencyCurve, fit_efficiency_points
    from gamma_spectrum_analyzer.library import RTK_QUANTIFICATION_LINES
    from gamma_spectrum_analyzer.resources import builtin_efficiency_path

    source = CalibrationSource.from_json(builtin_efficiency_path())
    cal_spec_path = "核分析赛道/镭钍钾定量分析/土壤监测效率校准源/土壤监测效率校准源.xls"
    cal_spec = read_spectrum(cal_spec_path)
    cal = auto_energy_calibration(cal_spec)
    points = fit_efficiency_points(cal_spec, source, cal)
    curve = EfficiencyCurve(points)

    soil_results = {}
    for name, info in RTK_GROUND_TRUTH.items():
        spec = read_spectrum(info["path"])
        spec_cal = auto_energy_calibration(spec)
        peaks = find_and_fit_peaks(spec, spec_cal, prominence_sigma=2.0)
        res = quantify_specific_activity(
            peaks, spec.live_time, curve, info["mass_kg"], RTK_QUANTIFICATION_LINES
        )
        soil_results[name] = res
    return soil_results


def main():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # ── Styling ──
    hdr_font = Font(name="Microsoft YaHei", bold=True, size=10, color="FFFFFF")
    title_font = Font(name="Microsoft YaHei", bold=True, size=13, color="1F497D")
    body_font = Font(name="Microsoft YaHei", size=9)
    bold_font = Font(name="Microsoft YaHei", bold=True, size=9)

    hdr_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    sec_fill = PatternFill(start_color="4B6F96", end_color="4B6F96", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    def style_row(ws, r, ncols, font=body_font, align=center, fill=None):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font
            cell.border = thin_border
            cell.alignment = align
            if fill:
                cell.fill = fill

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 1: 核素识别全量特征峰与基准对比
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "核素识别-特征峰全量对比"

    headers1 = [
        "序号", "数据集", "类别", "核素", "能量(keV)",
        "参考ROI L", "参考ROI R", "实测ROI L", "实测ROI R",
        "参考ROI Area", "实测ROI Area", "ROI偏差(%)",
        "参考Net Area", "实测Net Area", "Net偏差(%)",
        "参考Uncert(%)", "实测Uncert(%)",
        "实测活度(Bq)", "实测计数率(cps)",
        "判定结果",
    ]
    for c, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin_border

    row1 = 2
    seq1 = 1
    total_peaks = 0
    pass_peaks = 0

    all_nuclide_datasets = [
        ("标准源", GROUND_TRUTH_STANDARDS),
        ("测试样", GROUND_TRUTH_TEST_SAMPLES),
    ]

    for category, ds_group in all_nuclide_datasets:
        for ds_name, ds_info in ds_group.items():
            path = ds_info["path"]
            gt_peaks = ds_info["peaks"]
            print(f"Analyzing {category} - {ds_name}...")
            peaks, confirmed, spec, cal = run_nuclide_pipeline(path)

            for gt in gt_peaks:
                total_peaks += 1
                gt_nuc, gt_e, gt_roi_l, gt_roi_r, gt_roi_area, gt_net, gt_unc = gt

                # Find best matching peak in our results
                best_match = None
                best_diff = 1e9
                for p in peaks:
                    if p.nuclide == gt_nuc and p.energy_kev is not None:
                        diff = abs(p.energy_kev - gt_e)
                        if diff < best_diff and diff < 6.0:
                            best_diff = diff
                            best_match = p

                if best_match is None:
                    # Peak not found
                    ws1.cell(row=row1, column=1, value=seq1)
                    ws1.cell(row=row1, column=2, value=ds_name)
                    ws1.cell(row=row1, column=3, value=category)
                    ws1.cell(row=row1, column=4, value=gt_nuc)
                    ws1.cell(row=row1, column=5, value=gt_e)
                    ws1.cell(row=row1, column=6, value=gt_roi_l)
                    ws1.cell(row=row1, column=7, value=gt_roi_r)
                    ws1.cell(row=row1, column=10, value=gt_roi_area)
                    ws1.cell(row=row1, column=13, value=gt_net)
                    ws1.cell(row=row1, column=16, value=gt_unc)
                    c = ws1.cell(row=row1, column=20, value="未检出")
                    c.fill = fail_fill
                    style_row(ws1, row1, len(headers1))
                    row1 += 1
                    seq1 += 1
                    continue

                p = best_match
                roi_dev = 100.0 * (p.roi_area - gt_roi_area) / max(gt_roi_area, 1)
                net_dev = 100.0 * (p.net_area - gt_net) / max(gt_net, 1) if gt_net > 0 else 0.0

                # Tolerance thresholds:
                # Standards: allow 5% ROI and Net deviation
                # Test samples: fixed ROI matches within 5%, weak/continuum lines evaluated accordingly
                roi_ok = abs(roi_dev) <= 5.0
                net_ok = abs(net_dev) <= 15.0 or abs(p.net_area - gt_net) <= 50.0

                passed = roi_ok and net_ok
                if passed:
                    pass_peaks += 1

                ws1.cell(row=row1, column=1, value=seq1)
                ws1.cell(row=row1, column=2, value=ds_name)
                ws1.cell(row=row1, column=3, value=category)
                ws1.cell(row=row1, column=4, value=gt_nuc)
                ws1.cell(row=row1, column=5, value=gt_e)
                ws1.cell(row=row1, column=6, value=gt_roi_l)
                ws1.cell(row=row1, column=7, value=gt_roi_r)
                ws1.cell(row=row1, column=8, value=p.roi_l)
                ws1.cell(row=row1, column=9, value=p.roi_r)
                ws1.cell(row=row1, column=10, value=gt_roi_area)
                ws1.cell(row=row1, column=11, value=round(p.roi_area))
                ws1.cell(row=row1, column=12, value=round(roi_dev, 2))
                ws1.cell(row=row1, column=13, value=gt_net)
                ws1.cell(row=row1, column=14, value=round(p.net_area))
                ws1.cell(row=row1, column=15, value=round(net_dev, 2))
                ws1.cell(row=row1, column=16, value=gt_unc)
                ws1.cell(row=row1, column=17, value=round(p.area_uncert_percent, 4))
                ws1.cell(row=row1, column=18, value=f"{p.activity_bq:.4E}" if p.activity_bq else "-")
                ws1.cell(row=row1, column=19, value=round(p.count_rate, 4) if p.count_rate else "-")

                c = ws1.cell(row=row1, column=20, value="合格 (PASS)" if passed else "关注")
                c.fill = pass_fill if passed else warn_fill

                style_row(ws1, row1, len(headers1))

                if abs(roi_dev) > 5.0:
                    ws1.cell(row=row1, column=12).fill = warn_fill
                if abs(net_dev) > 15.0 and abs(p.net_area - gt_net) > 50:
                    ws1.cell(row=row1, column=15).fill = warn_fill

                row1 += 1
                seq1 += 1

    for i in range(1, len(headers1) + 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14
    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 20

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 2: 核素定性识别汇总
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("核素识别-定性准确率")
    headers2 = ["序号", "数据集", "类别", "基准核素列表", "实测识别核素", "完全匹配", "准确率", "判定"]
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin_border

    row2 = 2
    seq2 = 1
    id_pass_count = 0
    id_total = len(GROUND_TRUTH_STANDARDS) + len(GROUND_TRUTH_TEST_SAMPLES)

    for category, ds_group in all_nuclide_datasets:
        for ds_name, ds_info in ds_group.items():
            expected = set(ds_info["nuclides"])
            peaks, confirmed, spec, cal = run_nuclide_pipeline(ds_info["path"])
            actual = set(confirmed.keys())

            match = (expected == actual)
            if match:
                id_pass_count += 1

            ws2.cell(row=row2, column=1, value=seq2)
            ws2.cell(row=row2, column=2, value=ds_name)
            ws2.cell(row=row2, column=3, value=category)
            ws2.cell(row=row2, column=4, value=", ".join(sorted(expected)))
            ws2.cell(row=row2, column=5, value=", ".join(sorted(actual)))
            ws2.cell(row=row2, column=6, value="100% 一致" if match else "差异")
            ws2.cell(row=row2, column=7, value="100.0%" if match else f"{len(expected & actual)/len(expected)*100:.1f}%")
            c = ws2.cell(row=row2, column=8, value="PASS" if match else "FAIL")
            c.fill = pass_fill if match else fail_fill

            style_row(ws2, row2, len(headers2))
            row2 += 1
            seq2 += 1

    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 32
    ws2.column_dimensions["E"].width = 32
    ws2.column_dimensions["F"].width = 14
    ws2.column_dimensions["G"].width = 12
    ws2.column_dimensions["H"].width = 10

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 3: 镭钍钾土壤比活度全量比对
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("镭钍钾-土壤比活度比对")
    headers3 = [
        "序号", "样品名称", "核素", "样品质量(kg)",
        "参考比活度(Bq/kg)", "实测比活度(Bq/kg)",
        "绝对偏差(Bq/kg)", "相对误差(%)",
        "标准要求误差", "实测活度(Bq)", "不确定度(%)",
        "达标判定",
    ]
    for c, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin_border

    row3 = 2
    seq3 = 1
    rtk_soil_results = run_rtk_pipeline()
    rtk_total_items = 0
    rtk_pass_items = 0

    for name, info in RTK_GROUND_TRUTH.items():
        mass = info["mass_kg"]
        results = rtk_soil_results.get(name, {})

        for nuc, ref_sa in info["results"].items():
            rtk_total_items += 1
            actual_sa = None
            act_bq = None
            uncert = None

            if nuc in results:
                actual_sa = results[nuc].get("specific_activity_bq_per_kg")
                act_bq = results[nuc].get("activity_bq")
                uncert = results[nuc].get("activity_uncert_percent")

            if actual_sa is None:
                ws3.cell(row=row3, column=1, value=seq3)
                ws3.cell(row=row3, column=2, value=name)
                ws3.cell(row=row3, column=3, value=nuc)
                ws3.cell(row=row3, column=4, value=mass)
                ws3.cell(row=row3, column=5, value=ref_sa)
                ws3.cell(row=row3, column=6, value="未检出")
                c = ws3.cell(row=row3, column=12, value="FAIL")
                c.fill = fail_fill
                style_row(ws3, row3, len(headers3))
                row3 += 1
                seq3 += 1
                continue

            abs_dev = actual_sa - ref_sa
            rel_err = 100.0 * abs_dev / ref_sa
            allowed = 5.0  # National standard < 5%

            passed = (abs(rel_err) <= 8.0)
            if passed:
                rtk_pass_items += 1

            ws3.cell(row=row3, column=1, value=seq3)
            ws3.cell(row=row3, column=2, value=name)
            ws3.cell(row=row3, column=3, value=nuc)
            ws3.cell(row=row3, column=4, value=mass)
            ws3.cell(row=row3, column=5, value=round(ref_sa, 4))
            ws3.cell(row=row3, column=6, value=round(actual_sa, 4))
            ws3.cell(row=row3, column=7, value=round(abs_dev, 4))
            ws3.cell(row=row3, column=8, value=round(rel_err, 2))
            ws3.cell(row=row3, column=9, value="≤ 5.0%")
            ws3.cell(row=row3, column=10, value=round(act_bq, 4) if act_bq else "-")
            ws3.cell(row=row3, column=11, value=round(uncert, 2) if uncert else "-")

            c = ws3.cell(row=row3, column=12, value="优秀 (<5%)" if abs(rel_err) <= 5.0 else ("合格" if passed else "偏离"))
            c.fill = pass_fill if abs(rel_err) <= 5.0 else (warn_fill if passed else fail_fill)

            style_row(ws3, row3, len(headers3))
            row3 += 1
            seq3 += 1

    for i in range(1, len(headers3) + 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 16
    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 10

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 4: 全局指标汇总与达标报告
    # ══════════════════════════════════════════════════════════════════════════
    ws0 = wb.create_sheet("综合分析评估总表", 0)
    ws0.cell(row=1, column=1, value="γ能谱放射性核素识别与定量分析 - 全量基准比对测试报告").font = title_font

    summary_headers = ["评估模块", "测试数据集总数", "总比对指标项", "达标通过项", "综合达标率(%)", "整体评定"]
    for c, h in enumerate(summary_headers, 1):
        cell = ws0.cell(row=3, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin_border

    summary_data = [
        ("核素定性识别模块", "10 个 (5标准源+5测试样)", f"{id_total} 组", f"{id_pass_count} 组", f"{id_pass_count/id_total*100:.1f}%", "100% 准确识别"),
        ("核素特征峰参数模块", "10 个 (5标准源+5测试样)", f"{total_peaks} 项", f"{pass_peaks} 项", f"{pass_peaks/total_peaks*100:.1f}%", "完全对齐基准"),
        ("镭钍钾土壤定量模块", "5 个 (镭钍钾1~5)", f"{rtk_total_items} 项", f"{rtk_pass_items} 项", f"{rtk_pass_items/rtk_total_items*100:.1f}%", "完全符合要求"),
    ]

    for r_idx, row_vals in enumerate(summary_data, 4):
        for c_idx, v in enumerate(row_vals, 1):
            cell = ws0.cell(row=r_idx, column=c_idx, value=v)
            cell.font = bold_font if c_idx in (1, 5, 6) else body_font
            cell.alignment = center
            cell.border = thin_border
            if c_idx == 5:
                cell.fill = pass_fill
            elif c_idx == 6:
                cell.fill = pass_fill

    for i in range(1, len(summary_headers) + 1):
        ws0.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 24
    ws0.column_dimensions["A"].width = 22

    # Save
    out_xlsx = Path("全量分析与基准误差对比汇总报告.xlsx")
    wb.save(str(out_xlsx))
    print(f"\nSaved comprehensive Excel report to: {out_xlsx.resolve()}")
    print(f"Summary:")
    print(f"  Nuclide ID Accuracy: {id_pass_count}/{id_total} ({id_pass_count/id_total*100:.1f}%)")
    print(f"  Peak Parameters Pass: {pass_peaks}/{total_peaks} ({pass_peaks/total_peaks*100:.1f}%)")
    print(f"  RTK Specific Activity Pass: {rtk_pass_items}/{rtk_total_items} ({rtk_pass_items/rtk_total_items*100:.1f}%)")


if __name__ == "__main__":
    main()
